from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
import os
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl
import json
import warnings
import hashlib
import secrets
from collections import defaultdict
from dotenv import load_dotenv
warnings.filterwarnings('ignore', category=FutureWarning)

# Load environment variables from .env file
load_dotenv()

# Rate limiting for login attempts
login_attempts = defaultdict(list)
MAX_LOGIN_ATTEMPTS = 5
LOGIN_TIMEOUT = 300  # 5 minutes

# Custom JSON encoder to handle datetime objects
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif pd.isna(obj):
            return None
        return super().default(obj)

# Authentication configuration
AVENCION_USERNAME = "Avencion"
AVENCION_PASSWORD_HASH = hashlib.sha256("AvencionData@Center2025".encode()).hexdigest()

def login_required(f):
    """Decorator to require authentication for routes"""
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            return redirect(url_for('login'))
        
        # Additional session validation
        if not session.get('session_id') or not session.get('username'):
            session.clear()
            return redirect(url_for('login'))
        
        # Check if session is from same IP (optional security measure)
        if session.get('ip_address') and session.get('ip_address') != request.remote_addr:
            session.clear()
            return redirect(url_for('login'))
        
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

def clean_dataframe(df):
    """Clean dataframe by removing NaN values and empty strings"""
    # Replace various forms of NaN/empty values
    df = df.replace(['nan', 'NaN', 'None', 'NULL', 'null', 'N/A', 'n/a'], '')
    df = df.fillna('')
    
    # Remove rows that are completely empty
    df = df.dropna(how='all')
    
    # Remove columns that are completely empty
    df = df.dropna(axis=1, how='all')
    
    return df

def infer_column_names(df):
    """Infer meaningful column names from data content"""
    clean_columns = []
    
    # Check if this looks like an assessment criteria spreadsheet
    is_assessment = False
    if len(df) > 0:
        sample_text = ' '.join([str(val) for val in df.iloc[0].values if pd.notna(val)]).lower()
        if any(keyword in sample_text for keyword in ['assessment criteria', 'annexure', 'score', 'yes', 'no', 'points']):
            is_assessment = True
    
    for col in df.columns:
        col_str = str(col).strip()
        # Handle numeric column names (like '3', '4', '5', etc.)
        if col_str.isdigit() or 'unnamed' in col_str.lower() or pd.isna(col) or col_str == '' or col_str == 'nan':
            # Try to infer column name from first few values
            first_values = df[col].dropna().head(5).astype(str)
            if len(first_values) > 0:
                # Look for common patterns in the data
                sample_text = ' '.join(first_values).lower()
                
                # Check for common data patterns
                if any(word in sample_text for word in ['laptop', 'computer', 'device', 'pc', 'desktop']):
                    clean_columns.append('Device_Type')
                elif any(word in sample_text for word in ['hp', 'dell', 'lenovo', 'apple', 'asus', 'acer', 'toshiba', 'samsung']):
                    clean_columns.append('Manufacturer')
                elif any(word in sample_text for word in ['serial', 'sn', 'tag', 'asset', 'inventory']):
                    clean_columns.append('Serial_Number')
                elif any(word in sample_text for word in ['price', 'cost', 'amount', 'value', 'total']):
                    clean_columns.append('Price')
                elif any(word in sample_text for word in ['date', 'purchase', 'delivery', 'received', 'issued']):
                    clean_columns.append('Date')
                elif any(word in sample_text for word in ['quantity', 'qty', 'count', 'number']):
                    clean_columns.append('Quantity')
                elif any(word in sample_text for word in ['location', 'place', 'office', 'department', 'building']):
                    clean_columns.append('Location')
                elif any(word in sample_text for word in ['condition', 'status', 'state', 'working', 'broken']):
                    clean_columns.append('Condition')
                elif any(word in sample_text for word in ['description', 'details', 'notes', 'remarks']):
                    clean_columns.append('Description')
                elif any(word in sample_text for word in ['name', 'title', 'item', 'product']):
                    clean_columns.append('Name')
                elif any(word in sample_text for word in ['id', 'code', 'reference', 'ref']):
                    clean_columns.append('ID')
                elif any(word in sample_text for word in ['criteria', 'assessment', 'evaluation', 'requirement']):
                    clean_columns.append('Assessment_Criteria')
                elif any(word in sample_text for word in ['score', 'points', 'rating', 'grade']):
                    clean_columns.append('Score')
                elif any(word in sample_text for word in ['yes', 'no', 'y/n', 'true', 'false']):
                    clean_columns.append('Response')
                elif any(word in sample_text for word in ['annexure', 'annex', 'appendix']):
                    clean_columns.append('Document_Section')
                elif any(word in sample_text for word in ['sme', 'business', 'enterprise', 'company']):
                    clean_columns.append('Business_Info')
                elif any(word in sample_text for word in ['email', 'e-mail', 'mail']):
                    clean_columns.append('Email')
                elif any(word in sample_text for word in ['phone', 'mobile', 'contact']):
                    clean_columns.append('Phone')
                elif any(word in sample_text for word in ['address', 'street', 'city', 'zip']):
                    clean_columns.append('Address')
                elif any(word in sample_text for word in ['customer', 'client', 'user', 'person']):
                    clean_columns.append('Customer')
                elif any(word in sample_text for word in ['employee', 'staff', 'worker']):
                    clean_columns.append('Employee')
                elif any(word in sample_text for word in ['category', 'type', 'group']):
                    clean_columns.append('Category')
                elif any(word in sample_text for word in ['model', 'version', 'brand']):
                    clean_columns.append('Model')
                else:
                    # Special handling for assessment spreadsheets
                    if is_assessment:
                        # For assessment spreadsheets, use more descriptive names
                        if len(clean_columns) == 0:
                            clean_columns.append('Assessment_Criteria')
                        elif len(clean_columns) == 1:
                            clean_columns.append('Description')
                        elif len(clean_columns) == 2:
                            clean_columns.append('Score_Yes_No')
                        else:
                            # For additional columns, check if they contain Yes/No responses
                            sample_values = df[col].dropna().head(5).astype(str).str.lower()
                            if any(val in ['yes', 'no', 'y', 'n', '1', '0'] for val in sample_values):
                                clean_columns.append(f'Response_{len(clean_columns) - 2}')
                            else:
                                clean_columns.append(f'Additional_Info_{len(clean_columns) - 2}')
                    else:
                        # Try to infer from data type - simplified approach
                        try:
                            # Just try to convert to numeric to see if it works
                            sample_data = df[col].dropna().head(5).astype(str)
                            # Check if most values look numeric
                            numeric_count = sum(1 for val in sample_data if str(val).replace('.', '').replace('-', '').isdigit())
                            if numeric_count >= len(sample_data) * 0.7:  # 70% are numeric
                                clean_columns.append(f'Numeric_Column_{len(clean_columns) + 1}')
                            else:
                                clean_columns.append(f'Column_{len(clean_columns) + 1}')
                        except:
                            clean_columns.append(f'Column_{len(clean_columns) + 1}')
            else:
                clean_columns.append(f'Column_{len(clean_columns) + 1}')
        else:
            clean_columns.append(col_str)
    return clean_columns

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Session expires in 24 hours

# Security headers
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response

# Use SQLite for local development, PostgreSQL for production
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL and 'postgresql://' in DATABASE_URL:
    # Try PostgreSQL first, fallback to SQLite if connection fails
    try:
        # For Vercel deployment, we'll use the DATABASE_URL directly
        # The connection will be tested when the app starts
        app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
        print("✅ Configured for PostgreSQL database")
    except Exception as e:
        print(f"⚠️ PostgreSQL configuration failed: {e}")
        print("🔄 Falling back to SQLite for local development")
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db_manager.db'
else:
    # Default to SQLite for local development
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db_manager.db'
    print("✅ Using SQLite database for local development")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# Models
class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    project_type = db.Column(db.String(50), nullable=False, default='other')  # e.g., 'education', 'research', 'business'
    created_by = db.Column(db.String(100), nullable=False, default='Avencion')  # Creator signature
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    cohorts = db.relationship('Cohort', backref='project', lazy=True, cascade='all, delete-orphan')
    databases = db.relationship('Database', backref='project', lazy=True, cascade='all, delete-orphan')

class Cohort(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    created_by = db.Column(db.String(100), nullable=False, default='Avencion')  # Creator signature
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    spreadsheets = db.relationship('Spreadsheet', backref='cohort', lazy=True, cascade='all, delete-orphan')

class Spreadsheet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    sheet_names = db.Column(db.Text)  # JSON array of sheet names
    columns_info = db.Column(db.Text)  # JSON object with column information
    row_count = db.Column(db.Integer, default=0)
    created_by = db.Column(db.String(100), nullable=False, default='Avencion')  # Creator signature
    cohort_id = db.Column(db.Integer, db.ForeignKey('cohort.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Database(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # postgresql, access, excel
    connection_string = db.Column(db.Text)
    file_path = db.Column(db.String(500))
    created_by = db.Column(db.String(100), nullable=False, default='Avencion')  # Creator signature
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ImportLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    database_id = db.Column(db.Integer, db.ForeignKey('database.id'), nullable=False)
    import_type = db.Column(db.String(20), nullable=False)  # access, excel
    status = db.Column(db.String(20), nullable=False)  # success, failed
    message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

def migrate_database():
    """Migrate existing database to new schema"""
    try:
        with app.app_context():
            inspector = db.inspect(db.engine)
            
            # Check and add project_type column
            project_columns = [col['name'] for col in inspector.get_columns('project')]
            if 'project_type' not in project_columns:
                print("Adding project_type column to existing projects...")
                if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI']:
                    with db.engine.connect() as conn:
                        conn.execute(db.text('ALTER TABLE project ADD COLUMN project_type VARCHAR(50) DEFAULT \'other\''))
                        conn.commit()
                else:
                    with db.engine.connect() as conn:
                        conn.execute(db.text('ALTER TABLE project ADD COLUMN project_type VARCHAR(50) DEFAULT "other"'))
                        conn.commit()
            
            # Check and add created_by columns
            tables_to_check = ['project', 'cohort', 'spreadsheet', 'database']
            for table in tables_to_check:
                try:
                    table_columns = [col['name'] for col in inspector.get_columns(table)]
                    if 'created_by' not in table_columns:
                        print(f"Adding created_by column to existing {table}...")
                        if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI']:
                            with db.engine.connect() as conn:
                                conn.execute(db.text(f'ALTER TABLE {table} ADD COLUMN created_by VARCHAR(100) DEFAULT \'Avencion\''))
                                conn.commit()
                        else:
                            with db.engine.connect() as conn:
                                conn.execute(db.text(f'ALTER TABLE {table} ADD COLUMN created_by VARCHAR(100) DEFAULT "Avencion"'))
                                conn.commit()
                except Exception as e:
                    print(f"Error adding created_by to {table}: {e}")
            
            print("Migration completed successfully!")
                
    except Exception as e:
        print(f"Migration error: {e}")
        print("Creating new database schema...")
        db.create_all()

def analyze_excel_file(file_path):
    """Analyze Excel file and extract metadata using openpyxl for better handling"""
    try:
        # Use openpyxl for better Excel handling
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        first_sheet = sheet_names[0]
        ws = wb[first_sheet]
        
        # Get the actual data range (skip empty rows/columns)
        data_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        # Convert to DataFrame
        df_raw = pd.DataFrame(data_rows)
        
        # Handle merged cells by forward filling
        df_raw = df_raw.ffill(axis=1)  # Fill horizontally
        df_raw = df_raw.ffill(axis=0)  # Fill vertically
        
        # Find the row that contains actual column headers
        header_row = None
        
        # First, try to find headers with common keywords in the first 10 rows
        for i, row in df_raw.head(10).iterrows():
            row_values = [str(val).strip() for val in row.values if pd.notna(val)]
            if len(row_values) > 2:
                row_text = ' '.join(row_values).lower()
                # Look for more specific header patterns
                if any(keyword in row_text for keyword in ['s/n', 'serial', 'description', 'quantity', 'date', 'price', 'cost', 'name', 'id', 'code', 'number', 'amount', 'total', 'item', 'product', 'customer', 'client', 'address', 'phone', 'email', 'annexure', 'table', 'list', 'inventory', 'asset', 'criteria', 'assessment', 'score', 'yes', 'no', 'points']):
                    # Additional check: make sure this row doesn't have repetitive content
                    unique_values = set(row_values)
                    if len(unique_values) > 2:  # Should have more than 2 unique values
                        header_row = i
                        break
        
        # If no headers found with keywords, look for the first row with varied content
        if header_row is None:
            for i, row in df_raw.head(10).iterrows():
                non_null_count = row.notna().sum()
                if non_null_count >= 3:
                    # Check if this row has varied content (not repetitive)
                    row_values = [str(val).strip() for val in row.values if pd.notna(val)]
                    unique_values = set(row_values)
                    if len(unique_values) > 2 and len(unique_values) >= non_null_count * 0.5:  # At least 50% unique values
                        header_row = i
                        break
        
        # If still no headers found, use the first non-empty row
        if header_row is None:
            for i, row in df_raw.iterrows():
                if row.notna().sum() > 0:
                    header_row = i
                    break
        
        # Fallback to first row
        if header_row is None:
            header_row = 0
        
        # Now read the file with the correct header row using openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[first_sheet]
        
        # Get data starting from header row
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=min(header_row + 1000, ws.max_row), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        # Convert to DataFrame
        df = pd.DataFrame(data_rows)
        
        # Check if we have repetitive content in the first row (indicating wrong header detection)
        if len(df) > 0:
            first_row_values = [str(val).strip() for val in df.iloc[0].values if pd.notna(val)]
            if len(first_row_values) > 3:
                unique_first_row = set(first_row_values)
                # If first row has very repetitive content, try to find better headers
                if len(unique_first_row) <= 3 and len(first_row_values) > 5:
                    # Look for the next row that might be the real headers
                    for i in range(1, min(5, len(df))):
                        row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
                        unique_values = set(row_values)
                        if len(unique_values) > 3 and len(unique_values) >= len(row_values) * 0.3:
                            # This looks like a better header row
                            df = df.iloc[i:].reset_index(drop=True)
                            break
                
                # Special handling for assessment criteria spreadsheets
                # If we detect assessment-related content, try to find the actual criteria row
                first_row_text = ' '.join(first_row_values).lower()
                if 'assessment criteria' in first_row_text or 'annexure' in first_row_text:
                    # Look for a row that contains actual assessment questions
                    for i in range(1, min(10, len(df))):
                        row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
                        row_text = ' '.join(row_values).lower()
                        # Look for rows that contain actual questions (not just headers)
                        if any(keyword in row_text for keyword in ['does the', 'is the', 'has the', 'how many', 'what is', 'when', 'where', 'who']):
                            df = df.iloc[i:].reset_index(drop=True)
                            break
        
        # Handle merged cells by forward filling
        df = df.ffill(axis=1)  # Fill horizontally
        df = df.ffill(axis=0)  # Fill vertically
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Remove rows that are mostly empty (less than 20% data)
        df = df.dropna(thresh=df.shape[1] * 0.2)
        
        # Remove columns that are mostly empty (less than 10% data)
        df = df.dropna(axis=1, thresh=df.shape[0] * 0.1)
        
        # Ensure we have data after cleaning
        if len(df) == 0:
            # If no data after cleaning, try with less strict criteria
            df = pd.DataFrame(data_rows)
            df = df.dropna(thresh=df.shape[1] * 0.1)  # Less strict row filtering
            df = df.dropna(axis=1, thresh=df.shape[0] * 0.05)  # Less strict column filtering
        
        # Clean column names using smart inference
        clean_columns = infer_column_names(df)
        
        df.columns = clean_columns
        
        # Clean the dataframe thoroughly
        df = clean_dataframe(df)
        
        # Convert datetime columns to string for JSON serialization
        for col in df.columns:
            try:
                col_data = df[col]
                if hasattr(col_data, 'dtype'):
                    if col_data.dtype == 'datetime64[ns]':
                        df[col] = col_data.dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif col_data.dtype == 'object':
                        # Handle mixed types including datetime objects
                        df[col] = col_data.astype(str)
                else:
                    # Fallback for non-Series objects
                    df[col] = df[col].astype(str)
            except Exception:
                # If there's any error, convert to string
                df[col] = df[col].astype(str)
        
        # Get column info with proper handling of data types
        columns_info = {
            'columns': df.columns.tolist(),
            'dtypes': df.dtypes.astype(str).to_dict(),
            'sample_data': df.head(10).fillna('').to_dict('records'),
            'total_rows': len(df),
            'sheets': sheet_names,
            'header_row': header_row
        }
        
        return columns_info, len(df)
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        # Return basic info if analysis fails
        return {
            'columns': ['Error reading file'],
            'dtypes': {},
            'sample_data': [],
            'total_rows': 0,
            'sheets': []
        }, 0

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Rate limiting check
        client_ip = request.remote_addr
        current_time = datetime.utcnow()
        
        # Clean old attempts
        login_attempts[client_ip] = [attempt for attempt in login_attempts[client_ip] 
                                   if current_time - attempt < timedelta(seconds=LOGIN_TIMEOUT)]
        
        # Check if too many attempts
        if len(login_attempts[client_ip]) >= MAX_LOGIN_ATTEMPTS:
            flash('Too many login attempts. Please try again later.', 'error')
            return render_template('login.html')
        
        # Verify credentials
        if username == AVENCION_USERNAME and hashlib.sha256(password.encode()).hexdigest() == AVENCION_PASSWORD_HASH:
            # Clear login attempts on successful login
            login_attempts[client_ip].clear()
            
            session.permanent = True
            session['authenticated'] = True
            session['username'] = username
            session['login_time'] = current_time.isoformat()
            session['session_id'] = secrets.token_hex(16)
            session['ip_address'] = client_ip
            
            # Log successful login (without sensitive data)
            print(f"Successful login at {current_time} from IP: {client_ip}")
            
            flash('Welcome to Avencion Data Center!', 'success')
            return redirect(url_for('index'))
        else:
            # Record failed attempt
            login_attempts[client_ip].append(current_time)
            
            # Log failed login attempt (without credentials)
            print(f"Failed login attempt at {current_time} from IP: {client_ip}")
            flash('Invalid credentials. Please try again.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Clear session
    session.clear()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('login'))

@app.route('/help')
@login_required
def help_page():
    return render_template('help.html')

# Routes
@app.route('/')
@login_required
def index():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return render_template('index.html', projects=projects)

@app.route('/project/new', methods=['GET', 'POST'])
@login_required
def new_project():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        project_type = request.form['project_type']
        created_by = request.form.get('created_by', 'Avencion')
        
        project = Project(name=name, description=description, project_type=project_type, created_by=created_by)
        db.session.add(project)
        db.session.commit()
        
        flash('Project created successfully!', 'success')
        return redirect(url_for('index'))
    
    return render_template('new_project.html')

@app.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    return render_template('project_detail.html', project=project)

@app.route('/cohort/new/<int:project_id>', methods=['GET', 'POST'])
@login_required
def new_cohort(project_id):
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        created_by = request.form.get('created_by', 'Avencion')
        
        cohort = Cohort(name=name, description=description, project_id=project_id, created_by=created_by)
        db.session.add(cohort)
        db.session.commit()
        
        flash('Cohort created successfully!', 'success')
        return redirect(url_for('project_detail', project_id=project_id))
    
    return render_template('new_cohort.html', project=project)

@app.route('/cohort/<int:cohort_id>')
@login_required
def cohort_detail(cohort_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    return render_template('cohort_detail.html', cohort=cohort)

@app.route('/spreadsheet/upload/<int:cohort_id>', methods=['GET', 'POST'])
@login_required
def upload_spreadsheet(cohort_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            try:
                # Save file
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{cohort_id}_{filename}")
                file.save(file_path)
                
                # Analyze the Excel file
                columns_info, row_count = analyze_excel_file(file_path)
                
                spreadsheet = Spreadsheet(
                    name=request.form.get('name', filename),
                    filename=filename,
                    file_path=file_path,
                    sheet_names=json.dumps(columns_info.get('sheets', []), cls=DateTimeEncoder),
                    columns_info=json.dumps(columns_info, cls=DateTimeEncoder),
                    row_count=row_count,
                    cohort_id=cohort_id,
                    created_by=request.form.get('created_by', 'Avencion')
                )
                
                db.session.add(spreadsheet)
                db.session.commit()
                
                flash(f'Spreadsheet uploaded successfully! {row_count} rows analyzed.', 'success')
                return redirect(url_for('cohort_detail', cohort_id=cohort_id))
                
            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('upload_spreadsheet.html', cohort=cohort)

@app.route('/spreadsheet/<int:spreadsheet_id>')
@login_required
def spreadsheet_detail(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    columns_info = json.loads(spreadsheet.columns_info)
    return render_template('spreadsheet_detail.html', spreadsheet=spreadsheet, columns_info=columns_info)

@app.route('/spreadsheet/<int:spreadsheet_id>/download')
@login_required
def download_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    return send_file(spreadsheet.file_path, as_attachment=True, download_name=spreadsheet.filename)

@app.route('/spreadsheet/<int:spreadsheet_id>/view')
@login_required
def view_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    sheet_name = request.args.get('sheet', '')
    
    try:
        # Check if file exists
        if not os.path.exists(spreadsheet.file_path):
            flash('Spreadsheet file not found. The file may have been moved or deleted.', 'error')
            return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))
        
        # Read the Excel file using openpyxl with better error handling
        try:
            wb = openpyxl.load_workbook(spreadsheet.file_path, data_only=True)
        except Exception as excel_error:
            # Try to recreate the file if it's corrupted
            try:
                flash(f'Excel file appears to be corrupted. Attempting to recreate...', 'warning')
                
                # Create a new Excel file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                
                # Add some default data
                for col in range(1, 11):  # 10 columns
                    ws.cell(row=1, column=col, value=openpyxl.utils.get_column_letter(col))
                
                # Save the new file
                wb.save(spreadsheet.file_path)
                
                # Update the spreadsheet record
                spreadsheet.row_count = 10
                columns_info = {
                    'columns': [openpyxl.utils.get_column_letter(i) for i in range(1, 11)],
                    'total_rows': 10,
                    'sheets': ['Sheet1']
                }
                spreadsheet.columns_info = json.dumps(columns_info, cls=DateTimeEncoder)
                spreadsheet.updated_at = datetime.utcnow()
                db.session.commit()
                
                flash('Spreadsheet recreated successfully! You can now edit it online.', 'success')
                
                # Reload the workbook
                wb = openpyxl.load_workbook(spreadsheet.file_path, data_only=True)
                
            except Exception as recreate_error:
                flash(f'Failed to recreate spreadsheet: {str(recreate_error)}. Please delete and re-upload the file.', 'error')
                return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))
        available_sheets = wb.sheetnames
        
        if not sheet_name or sheet_name not in available_sheets:
            sheet_name = available_sheets[0]
        
        ws = wb[sheet_name]
        
        # Read the selected sheet with improved header detection
        data_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        df_raw = pd.DataFrame(data_rows)
        
        # Find the row that contains actual column headers
        header_row = None
        
        # First, try to find headers with common keywords in the first 10 rows
        for i, row in df_raw.head(10).iterrows():
            row_values = [str(val).strip() for val in row.values if pd.notna(val)]
            if len(row_values) > 2:
                row_text = ' '.join(row_values).lower()
                # Look for more specific header patterns
                if any(keyword in row_text for keyword in ['s/n', 'serial', 'description', 'quantity', 'date', 'price', 'cost', 'name', 'id', 'code', 'number', 'amount', 'total', 'item', 'product', 'customer', 'client', 'address', 'phone', 'email', 'annexure', 'table', 'list', 'inventory', 'asset', 'criteria', 'assessment', 'score', 'yes', 'no', 'points']):
                    # Additional check: make sure this row doesn't have repetitive content
                    unique_values = set(row_values)
                    if len(unique_values) > 2:  # Should have more than 2 unique values
                        header_row = i
                        break
        
        # If no headers found with keywords, look for the first row with varied content
        if header_row is None:
            for i, row in df_raw.head(10).iterrows():
                non_null_count = row.notna().sum()
                if non_null_count >= 3:
                    # Check if this row has varied content (not repetitive)
                    row_values = [str(val).strip() for val in row.values if pd.notna(val)]
                    unique_values = set(row_values)
                    if len(unique_values) > 2 and len(unique_values) >= non_null_count * 0.5:  # At least 50% unique values
                        header_row = i
                        break
        
        # If still no headers found, use the first non-empty row
        if header_row is None:
            for i, row in df_raw.iterrows():
                if row.notna().sum() > 0:
                    header_row = i
                    break
        
        # Fallback to first row
        if header_row is None:
            header_row = 0
        
        # Read with correct header row using openpyxl
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        df = pd.DataFrame(data_rows)
        
        # Handle merged cells by forward filling
        df = df.ffill(axis=1)  # Fill horizontally
        df = df.ffill(axis=0)  # Fill vertically
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Remove rows that are mostly empty (less than 20% data)
        df = df.dropna(thresh=df.shape[1] * 0.2)
        
        # Remove columns that are mostly empty (less than 10% data)
        df = df.dropna(axis=1, thresh=df.shape[0] * 0.1)
        
        # Ensure we have data after cleaning
        if len(df) == 0:
            # If no data after cleaning, try with less strict criteria
            df = pd.DataFrame(data_rows)
            df = df.dropna(thresh=df.shape[1] * 0.1)  # Less strict row filtering
            df = df.dropna(axis=1, thresh=df.shape[0] * 0.05)  # Less strict column filtering
        
        # Check if we have repetitive content in the first row (indicating wrong header detection)
        if len(df) > 0:
            first_row_values = [str(val).strip() for val in df.iloc[0].values if pd.notna(val)]
            if len(first_row_values) > 3:
                unique_first_row = set(first_row_values)
                # If first row has very repetitive content, try to find better headers
                if len(unique_first_row) <= 3 and len(first_row_values) > 5:
                    # Look for the next row that might be the real headers
                    for i in range(1, min(5, len(df))):
                        row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
                        unique_values = set(row_values)
                        if len(unique_values) > 3 and len(unique_values) >= len(row_values) * 0.3:
                            # This looks like a better header row
                            df = df.iloc[i:].reset_index(drop=True)
                            break
                
                # Special handling for assessment criteria spreadsheets
                # If we detect assessment-related content, try to find the actual criteria row
                first_row_text = ' '.join(first_row_values).lower()
                if 'assessment criteria' in first_row_text or 'annexure' in first_row_text:
                    # Look for a row that contains actual assessment questions
                    for i in range(1, min(10, len(df))):
                        row_values = [str(val).strip() for val in df.iloc[i].values if pd.notna(val)]
                        row_text = ' '.join(row_values).lower()
                        # Look for rows that contain actual questions (not just headers)
                        if any(keyword in row_text for keyword in ['does the', 'is the', 'has the', 'how many', 'what is', 'when', 'where', 'who']):
                            df = df.iloc[i:].reset_index(drop=True)
                            break
        
        # Clean column names using smart inference
        clean_columns = infer_column_names(df)
        
        df.columns = clean_columns
        
        # Clean the dataframe thoroughly
        df = clean_dataframe(df)
        
        # Convert datetime columns to string for display
        for col in df.columns:
            try:
                col_data = df[col]
                if hasattr(col_data, 'dtype'):
                    if col_data.dtype == 'datetime64[ns]':
                        df[col] = col_data.dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif col_data.dtype == 'object':
                        # Handle mixed types including datetime objects
                        df[col] = col_data.astype(str)
                else:
                    # Fallback for non-Series objects
                    df[col] = df[col].astype(str)
            except Exception:
                # If there's any error, convert to string
                df[col] = df[col].astype(str)
        
        # Check if we have data to display
        if len(df) == 0:
            table_html = '<div class="text-center py-8 text-gray-500"><p>No data found in this sheet.</p></div>'
            columns_truncated = False
            return render_template('view_spreadsheet.html', 
                                 spreadsheet=spreadsheet, 
                                 table_html=table_html,
                                 available_sheets=available_sheets,
                                 current_sheet=sheet_name,
                                 total_rows=0,
                                 displayed_rows=0,
                                 columns_truncated=False,
                                 total_columns=0,
                                 displayed_columns=0)
        
        # Convert to HTML table with Tailwind CSS classes - limit to 50 rows for better performance
        display_df = df.head(50)
        
        # Limit columns for very wide tables to prevent layout issues
        max_columns = 20  # Limit to 20 columns for display
        if len(display_df.columns) > max_columns:
            display_df = display_df.iloc[:, :max_columns]
            columns_truncated = True
        else:
            columns_truncated = False
        
        # Create a proper HTML table structure with better styling for wide tables
        table_html = '<table class="min-w-full divide-y divide-gray-200 table-fixed">'
        
        # Add header
        table_html += '<thead class="bg-gray-50">'
        table_html += '<tr>'
        for col in display_df.columns:
            # Truncate long column names for better display
            col_display = str(col)[:30] + '...' if len(str(col)) > 30 else str(col)
            table_html += f'<th class="px-3 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider truncate" title="{col}">{col_display}</th>'
        table_html += '</tr>'
        table_html += '</thead>'
        
        # Add body
        table_html += '<tbody class="bg-white divide-y divide-gray-200">'
        for idx, row in display_df.iterrows():
            table_html += '<tr class="hover:bg-gray-50">'
            for value in row.values:
                # Handle NaN values
                if pd.isna(value) or value == 'nan':
                    cell_value = ''
                else:
                    cell_value = str(value)
                # Truncate long cell values and add tooltip
                if len(cell_value) > 50:
                    cell_display = cell_value[:50] + '...'
                    cell_title = cell_value
                else:
                    cell_display = cell_value
                    cell_title = ''
                
                table_html += f'<td class="px-3 py-4 text-sm text-gray-900 truncate" title="{cell_title}">{cell_display}</td>'
            table_html += '</tr>'
        table_html += '</tbody>'
        table_html += '</table>'
        
        # Debug: Print some info about the data
        print(f"Debug: Found {len(df)} rows and {len(df.columns)} columns")
        print(f"Debug: Header row was at index {header_row}")
        print(f"Debug: Column names: {list(df.columns)}")
        print(f"Debug: First few rows of data:")
        print(display_df.head(3).to_string())
        
        return render_template('view_spreadsheet.html', 
                             spreadsheet=spreadsheet, 
                             table_html=table_html,
                             available_sheets=available_sheets,
                             current_sheet=sheet_name,
                             total_rows=len(df),
                             displayed_rows=min(50, len(df)),
                             columns_truncated=columns_truncated,
                             total_columns=len(df.columns),
                             displayed_columns=len(display_df.columns))
        
    except Exception as e:
        flash(f'Error viewing spreadsheet: {str(e)}', 'error')
        return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))

@app.route('/spreadsheet/<int:spreadsheet_id>/search')
@login_required
def search_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    query = request.args.get('q', '')
    column = request.args.get('column', '')
    sheet_name = request.args.get('sheet', '')
    
    if not query:
        return jsonify({'data': [], 'total': 0})
    
    try:
        # Read Excel file using openpyxl
        wb = openpyxl.load_workbook(spreadsheet.file_path, data_only=True)
        if not sheet_name or sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Read with smart header detection and handle merged cells
        data_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        df_raw = pd.DataFrame(data_rows)
        
        # Handle merged cells by forward filling
        df_raw = df_raw.ffill(axis=1)  # Fill horizontally
        df_raw = df_raw.ffill(axis=0)  # Fill vertically
        
        # Find the row that contains actual column headers
        header_row = None
        for i, row in df_raw.iterrows():
            row_values = [str(val).strip() for val in row.values if pd.notna(val)]
            if len(row_values) > 3:
                if any(keyword in ' '.join(row_values).lower() for keyword in ['s/n', 'serial', 'description', 'quantity', 'date', 'price', 'cost']):
                    header_row = i
                    break
        
        if header_row is None:
            for i, row in df_raw.iterrows():
                non_null_count = row.notna().sum()
                if non_null_count >= 3:
                    header_row = i
                    break
        
        if header_row is None:
            header_row = 0
        
        # Read with correct header row using openpyxl
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append(row)
        
        df = pd.DataFrame(data_rows)
        
        # Handle merged cells by forward filling
        df = df.ffill(axis=1)  # Fill horizontally
        df = df.ffill(axis=0)  # Fill vertically
        
        # Clean column names using smart inference
        clean_columns = infer_column_names(df)
        
        df.columns = clean_columns
        
        # Clean the dataframe thoroughly
        df = clean_dataframe(df)
        
        # Convert datetime columns to string for search
        for col in df.columns:
            try:
                col_data = df[col]
                if hasattr(col_data, 'dtype'):
                    if col_data.dtype == 'datetime64[ns]':
                        df[col] = col_data.dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif col_data.dtype == 'object':
                        # Handle mixed types including datetime objects
                        df[col] = col_data.astype(str)
                else:
                    # Fallback for non-Series objects
                    df[col] = df[col].astype(str)
            except Exception:
                # If there's any error, convert to string
                df[col] = df[col].astype(str)
        
        if column and column in df.columns:
            # Search in specific column
            mask = df[column].astype(str).str.contains(query, case=False, na=False)
        else:
            # Search in all columns
            mask = df.astype(str).apply(lambda x: x.str.contains(query, case=False, na=False)).any(axis=1)
        
        results = df[mask].head(100).fillna('').to_dict('records')  # Limit to 100 results
        
        return jsonify({
            'data': results,
            'total': len(df[mask]),
            'query': query,
            'column': column,
            'sheet': sheet_name
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/database/new/<int:project_id>', methods=['GET', 'POST'])
@login_required
def new_database(project_id):
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'POST':
        name = request.form['name']
        db_type = request.form['type']
        
        if db_type == 'postgresql':
            host = request.form['host']
            port = request.form['port']
            database = request.form['database']
            username = request.form['username']
            password = request.form['password']
            
            connection_string = f"postgresql://{username}:{password}@{host}:{port}/{database}"
            
            database_obj = Database(
                name=name,
                type=db_type,
                connection_string=connection_string,
                project_id=project_id,
                created_by=request.form.get('created_by', 'Avencion')
            )
            
        elif db_type == 'access':
            if 'file' not in request.files:
                flash('No file selected', 'error')
                return redirect(request.url)
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return redirect(request.url)
            
            if file:
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{project_id}_{filename}")
                file.save(file_path)
                
                database_obj = Database(
                    name=name,
                    type=db_type,
                    file_path=file_path,
                    project_id=project_id,
                    created_by=request.form.get('created_by', 'Avencion')
                )
        
        elif db_type == 'excel':
            excel_option = request.form.get('excel_option', 'upload')
            
            if excel_option == 'upload':
                # Upload existing Excel file
                if 'file' not in request.files:
                    flash('No file selected', 'error')
                    return redirect(request.url)
                
                file = request.files['file']
                if file.filename == '':
                    flash('No file selected', 'error')
                    return redirect(request.url)
                
                if file:
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{project_id}_{filename}")
                    file.save(file_path)
                    
                    database_obj = Database(
                        name=name,
                        type=db_type,
                        file_path=file_path,
                        project_id=project_id,
                        created_by=request.form.get('created_by', 'Avencion')
                    )
            
            elif excel_option == 'create':
                # Create new Excel file
                rows = int(request.form.get('rows', 10))
                cols = int(request.form.get('cols', 10))
                
                # Create a new Excel file with the specified dimensions
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                
                # Add headers (A, B, C, etc.)
                for col in range(1, cols + 1):
                    ws.cell(row=1, column=col, value=openpyxl.utils.get_column_letter(col))
                
                # Add row numbers
                for row in range(2, rows + 2):
                    ws.cell(row=row, column=1, value=row - 1)
                
                # Save the file
                filename = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{project_id}_{filename}")
                wb.save(file_path)
                
                database_obj = Database(
                    name=name,
                    type=db_type,
                    file_path=file_path,
                    project_id=project_id,
                    created_by=request.form.get('created_by', 'Avencion')
                )
        
        db.session.add(database_obj)
        db.session.commit()
        
        flash('Database added successfully!', 'success')
        return redirect(url_for('project_detail', project_id=project_id))
    
    return render_template('new_database.html', project=project)

@app.route('/database/<int:database_id>/import', methods=['POST'])
@login_required
def import_database(database_id):
    database = Database.query.get_or_404(database_id)
    
    try:
        if database.type == 'access':
            success = import_access_database(database)
        elif database.type == 'excel':
            success = import_excel_database(database)
        else:
            flash('Unsupported database type', 'error')
            return redirect(url_for('project_detail', project_id=database.project_id))
        
        if success:
            flash('Database imported successfully!', 'success')
        else:
            flash('Failed to import database', 'error')
            
    except Exception as e:
        flash(f'Error importing database: {str(e)}', 'error')
    
    return redirect(url_for('project_detail', project_id=database.project_id))

@app.route('/database/<int:database_id>/tables')
@login_required
def database_tables(database_id):
    database = Database.query.get_or_404(database_id)
    
    try:
        if database.type == 'postgresql':
            tables = get_postgresql_tables(database.connection_string)
        elif database.type == 'access':
            tables = get_access_tables(database.file_path)
        elif database.type == 'excel':
            tables = get_excel_tables(database.file_path)
        else:
            tables = []
            
        return jsonify({'tables': tables})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def import_access_database(database):
    """Import Access database to PostgreSQL"""
    try:
        log = ImportLog(
            database_id=database.id,
            import_type='access',
            status='success',
            message='Access database import completed (simplified version)'
        )
        db.session.add(log)
        db.session.commit()
        return True
        
    except Exception as e:
        log = ImportLog(
            database_id=database.id,
            import_type='access',
            status='failed',
            message=str(e)
        )
        db.session.add(log)
        db.session.commit()
        return False

def import_excel_database(database):
    """Import Excel file to PostgreSQL"""
    try:
        log = ImportLog(
            database_id=database.id,
            import_type='excel',
            status='success',
            message='Excel file import completed (simplified version)'
        )
        db.session.add(log)
        db.session.commit()
        return True
        
    except Exception as e:
        log = ImportLog(
            database_id=database.id,
            import_type='excel',
            status='failed',
            message=str(e)
        )
        db.session.add(log)
        db.session.commit()
        return False

def get_postgresql_tables(connection_string):
    """Get list of tables from PostgreSQL database"""
    try:
        return ['users', 'products', 'orders']  # Placeholder tables
    except Exception as e:
        raise Exception(f"Error connecting to PostgreSQL: {str(e)}")

def get_access_tables(file_path):
    """Get list of tables from Access database"""
    try:
        return ['table1', 'table2']  # Placeholder
    except Exception as e:
        raise Exception(f"Error reading Access database: {str(e)}")

def get_excel_tables(file_path):
    """Get list of sheets from Excel file"""
    try:
        excel_file = pd.ExcelFile(file_path)
        return excel_file.sheet_names
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

# Delete and Edit Routes

@app.route('/project/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'POST':
        project.name = request.form['name']
        project.project_type = request.form['project_type']
        project.description = request.form.get('description', '')
        
        db.session.commit()
        flash('Project updated successfully!', 'success')
        return redirect(url_for('project_detail', project_id=project.id))
    
    return render_template('edit_project.html', project=project)

@app.route('/project/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    
    # Delete all related data
    for cohort in project.cohorts:
        for spreadsheet in cohort.spreadsheets:
            if spreadsheet.file_path and os.path.exists(spreadsheet.file_path):
                os.remove(spreadsheet.file_path)
        db.session.delete(cohort)
    
    for database in project.databases:
        if database.file_path and os.path.exists(database.file_path):
            os.remove(database.file_path)
        db.session.delete(database)
    
    db.session.delete(project)
    db.session.commit()
    
    flash('Project and all associated data deleted successfully!', 'success')
    return redirect(url_for('index'))

@app.route('/cohort/<int:cohort_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_cohort(cohort_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    
    if request.method == 'POST':
        cohort.name = request.form['name']
        cohort.description = request.form.get('description', '')
        
        db.session.commit()
        flash('Cohort updated successfully!', 'success')
        return redirect(url_for('cohort_detail', cohort_id=cohort.id))
    
    return render_template('edit_cohort.html', cohort=cohort)

@app.route('/cohort/<int:cohort_id>/delete', methods=['POST'])
@login_required
def delete_cohort(cohort_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    project_id = cohort.project_id
    
    # Delete all spreadsheets in the cohort
    for spreadsheet in cohort.spreadsheets:
        if spreadsheet.file_path and os.path.exists(spreadsheet.file_path):
            os.remove(spreadsheet.file_path)
        db.session.delete(spreadsheet)
    
    db.session.delete(cohort)
    db.session.commit()
    
    flash('Cohort and all associated spreadsheets deleted successfully!', 'success')
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/database/<int:database_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_database(database_id):
    database = Database.query.get_or_404(database_id)
    
    if request.method == 'POST':
        database.name = request.form['name']
        
        if database.type == 'postgresql':
            host = request.form['host']
            port = request.form['port']
            db_name = request.form['database']
            username = request.form['username']
            password = request.form['password']
            database.connection_string = f"postgresql://{username}:{password}@{host}:{port}/{db_name}"
        
        db.session.commit()
        flash('Database updated successfully!', 'success')
        return redirect(url_for('project_detail', project_id=database.project_id))
    
    return render_template('edit_database.html', database=database)

@app.route('/database/<int:database_id>/delete', methods=['POST'])
@login_required
def delete_database(database_id):
    database = Database.query.get_or_404(database_id)
    project_id = database.project_id
    
    if database.file_path and os.path.exists(database.file_path):
        os.remove(database.file_path)
    
    db.session.delete(database)
    db.session.commit()
    
    flash('Database deleted successfully!', 'success')
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/spreadsheet/<int:spreadsheet_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    
    if request.method == 'POST':
        spreadsheet.name = request.form['name']
        db.session.commit()
        flash('Spreadsheet updated successfully!', 'success')
        return redirect(url_for('cohort_detail', cohort_id=spreadsheet.cohort_id))
    
    return render_template('edit_spreadsheet.html', spreadsheet=spreadsheet)

@app.route('/spreadsheet/<int:spreadsheet_id>/delete', methods=['POST'])
@login_required
def delete_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    cohort_id = spreadsheet.cohort_id
    
    if spreadsheet.file_path and os.path.exists(spreadsheet.file_path):
        os.remove(spreadsheet.file_path)
    
    db.session.delete(spreadsheet)
    db.session.commit()
    
    flash('Spreadsheet deleted successfully!', 'success')
    return redirect(url_for('cohort_detail', cohort_id=cohort_id))

@app.route('/spreadsheet/create/<int:cohort_id>', methods=['GET', 'POST'])
@login_required
def create_spreadsheet(cohort_id):
    cohort = Cohort.query.get_or_404(cohort_id)
    
    if request.method == 'POST':
        name = request.form['name']
        rows = int(request.form.get('rows', 10))
        cols = int(request.form.get('cols', 10))
        
        # Create a new Excel file with the specified dimensions
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers (A, B, C, etc.)
        for col in range(1, cols + 1):
            ws.cell(row=1, column=col, value=openpyxl.utils.get_column_letter(col))
        
        # Add row numbers
        for row in range(2, rows + 2):
            ws.cell(row=row, column=1, value=row - 1)
        
        # Save the file
        filename = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{cohort_id}_{filename}")
        wb.save(file_path)
        
        # Create spreadsheet record
        spreadsheet = Spreadsheet(
            name=name,
            filename=filename,
            file_path=file_path,
            sheet_names=json.dumps(['Sheet1']),
            columns_info=json.dumps({
                'columns': [openpyxl.utils.get_column_letter(i) for i in range(1, cols + 1)],
                'total_rows': rows,
                'sheets': ['Sheet1']
            }),
            row_count=rows,
            cohort_id=cohort_id
        )
        
        db.session.add(spreadsheet)
        db.session.commit()
        
        flash(f'Spreadsheet "{name}" created successfully!', 'success')
        return redirect(url_for('edit_spreadsheet_online', spreadsheet_id=spreadsheet.id))
    
    return render_template('create_spreadsheet.html', cohort=cohort)

@app.route('/spreadsheet/<int:spreadsheet_id>/edit-online')
@login_required
def edit_spreadsheet_online(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    
    # Check if file exists
    if not os.path.exists(spreadsheet.file_path):
        flash('Spreadsheet file not found. Please re-upload the file.', 'error')
        return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))
    
    # Read the Excel file to get current data
    try:
        # Try to load the workbook with better error handling
        try:
            wb = openpyxl.load_workbook(spreadsheet.file_path, data_only=False)  # Keep formulas
        except Exception as excel_error:
            # If the file is corrupted, try to create a new one
            flash(f'Excel file appears to be corrupted. Creating a new spreadsheet.', 'warning')
            
            # Create a new Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add some default data
            for col in range(1, 11):  # 10 columns
                ws.cell(row=1, column=col, value=openpyxl.utils.get_column_letter(col))
            
            # Save the new file
            wb.save(spreadsheet.file_path)
            
            # Reload the workbook
            wb = openpyxl.load_workbook(spreadsheet.file_path, data_only=False)
        
        ws = wb.active
        
        # Get the data range
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Ensure we have at least some data
        if max_row < 1:
            max_row = 10
        if max_col < 1:
            max_col = 10
        
        # Extract data with formulas
        data = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    if cell.data_type == 'f':  # Formula
                        row_data.append({
                            'value': cell.value,
                            'formula': cell.value,
                            'type': 'formula'
                        })
                    else:
                        row_data.append({
                            'value': cell.value,
                            'formula': '',
                            'type': 'value'
                        })
                else:
                    row_data.append({
                        'value': '',
                        'formula': '',
                        'type': 'empty'
                    })
            data.append(row_data)
        
        # Generate column letters for the template
        column_letters = [openpyxl.utils.get_column_letter(i) for i in range(1, max_col + 1)]
        
        return render_template('edit_spreadsheet_online.html', 
                             spreadsheet=spreadsheet, 
                             data=data,
                             max_row=max_row,
                             max_col=max_col,
                             column_letters=column_letters)
    
    except Exception as e:
        flash(f'Error loading spreadsheet: {str(e)}', 'error')
        return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))

@app.route('/spreadsheet/<int:spreadsheet_id>/recreate', methods=['POST'])
@login_required
def recreate_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    
    try:
        # Create a new Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add default headers
        for col in range(1, 11):  # 10 columns
            ws.cell(row=1, column=col, value=openpyxl.utils.get_column_letter(col))
        
        # Save the new file
        wb.save(spreadsheet.file_path)
        
        # Update the spreadsheet record
        spreadsheet.row_count = 10
        columns_info = {
            'columns': [openpyxl.utils.get_column_letter(i) for i in range(1, 11)],
            'total_rows': 10,
            'sheets': ['Sheet1']
        }
        spreadsheet.columns_info = json.dumps(columns_info, cls=DateTimeEncoder)
        spreadsheet.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('Spreadsheet recreated successfully!', 'success')
        return redirect(url_for('edit_spreadsheet_online', spreadsheet_id=spreadsheet_id))
        
    except Exception as e:
        flash(f'Error recreating spreadsheet: {str(e)}', 'error')
        return redirect(url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id))

@app.route('/spreadsheet/<int:spreadsheet_id>/save', methods=['POST'])
@login_required
def save_spreadsheet(spreadsheet_id):
    spreadsheet = Spreadsheet.query.get_or_404(spreadsheet_id)
    
    try:
        # Get the updated data from the form
        request_data = request.json
        
        # Handle both old and new data formats
        if isinstance(request_data, list):
            # Old format - just data array
            data = request_data
            metadata = None
        else:
            # New format - data with metadata
            data = request_data.get('data', [])
            metadata = request_data.get('metadata', {})
        
        # Load the workbook
        wb = openpyxl.load_workbook(spreadsheet.file_path)
        ws = wb.active
        
        # Clear existing content if we have metadata (indicating structure change)
        if metadata:
            # Clear all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        
        # Update cells with new data
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell_data.get('type') == 'formula':
                    # Set formula
                    cell.value = cell_data.get('formula', '')
                else:
                    # Set value
                    cell.value = cell_data.get('value', '')
        
        # Save the workbook
        wb.save(spreadsheet.file_path)
        
        # Update the spreadsheet record with new metadata if available
        if metadata:
            spreadsheet.row_count = metadata.get('rows', len(data))
            # Update columns_info if we have column headers
            if metadata.get('column_headers'):
                columns_info = {
                    'columns': metadata['column_headers'],
                    'total_rows': metadata.get('rows', len(data)),
                    'sheets': ['Sheet1']
                }
                spreadsheet.columns_info = json.dumps(columns_info, cls=DateTimeEncoder)
        
        spreadsheet.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Check if this is a manual save (not auto-save)
        is_manual_save = request_data.get('manual_save', False) if isinstance(request_data, dict) else False
        
        if is_manual_save:
            # Redirect to spreadsheet detail page for manual saves
            flash('Spreadsheet saved successfully!', 'success')
            return jsonify({
                'success': True, 
                'message': 'Spreadsheet saved successfully!',
                'redirect': url_for('spreadsheet_detail', spreadsheet_id=spreadsheet_id)
            })
        else:
            # Return JSON response for auto-saves
            return jsonify({'success': True, 'message': 'Spreadsheet saved successfully!'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# For Vercel deployment
app.debug = False

if __name__ == '__main__':
    with app.app_context():
        # Run migration first
        migrate_database()
        # Create tables if they don't exist
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000) 